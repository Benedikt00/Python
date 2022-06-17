from playwright.sync_api import sync_playwright
from time import sleep
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook, load_workbook
import time



for x in range(20):
    with sync_playwright() as p:
        browser = p.chromium.launch()
        page = browser.new_page()
        page.goto('https://www.solarweb.com/PvSystems/PvSystem?pvSystemId=673dc46c-2863-4996-a638-a2480124484c')
        page.fill('input#username', 'b.simbuerger@aon.at')
        page.fill('input#password', 'HTL@Benedikt1')
        page.click('button[type=button]')
        for i in range(5):
            print("Progress " + str(i))
            sleep(1)

        html = page.content()
        soup = BeautifulSoup(html, 'html.parser')
        var = soup.find('span', {'style':'white-space: nowrap; font-weight: 500;'}).text
        print(var)

        wb = load_workbook('data.xlsx')
        ws = wb.active

        ws['A'+ str(x + 1)]= var

        ws['B' + str(x + 1)] = time.time()

        time_object = time.localtime()
        local_date = time.strftime("%y %m %d", time_object)
        ws['C'+ str(x + 1)]= local_date


        local_time = time.strftime("%H : %M: %S ", time_object)
        ws['D' + str(x + 1)] = local_time

        wb.save('data.xlsx')





        sleep(300)





