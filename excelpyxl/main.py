from openpyxl import Workbook, load_workbook

wb = load_workbook('Grades.xlsx')

ws = wb.active
ws['A2'].value = "Test"
print(ws['A2'])




